"""Swing 宽表扫描入口: 全池逐股 pattern 特征 → Excel

用法:
  python run_swing.py                  # 截至今天
  python run_swing.py --asof 2026-07-20
"""
from __future__ import annotations
import sys, io, argparse, datetime as dt, time
from pathlib import Path
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import yaml
import pandas as pd

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / "src"))

import provider as P
import swing as sw

OUT_XLSX = "HK_Swing_Pattern.xlsx"  # 固定下载文件名


COL_ORDER = [
    "Ticker", "Name", "Close",
    "周J",
    "周KDJ背离", "周MACD背离", "周RSI背离", "日MACD背离", "日RSI背离",
    "周度背离", "日度背离",
    "周度DeMark", "日度DeMark",
    "十字星(3d)", "涨放量跌缩量", "climax",
    "KDJ Cross", "日MACD Cross", "5_10 Cross",
]
# HK 专属 4 列 (年线斜率 / RS_HSI / RS_HSI_1W / RS_HSI_4W)
COL_ORDER_HK = COL_ORDER + ["年线斜率", "MRS_HSI", "MRS_HSI_1W", "MRS_HSI_4W",
                            "MRS_HD", "MRS_HD_1W", "MRS_HD_4W"]


def to_bloomberg(wind_code: str) -> str:
    num, suf = wind_code.split(".")
    return f"{num.lstrip('0') or num} {suf}"


def write_excel(df: pd.DataFrame, asof: str, out: Path, col_order: list = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active; ws.title = "swing"
    head = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = col_order or COL_ORDER
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c); cell.fill = head; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for i, (_, r) in enumerate(df.iterrows(), 2):
        for j, c in enumerate(cols, 1):
            v = r.get(c)
            if isinstance(v, float) and pd.isna(v):
                v = None
            cell = ws.cell(i, j, v)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    widths = {"Ticker": 9, "Name": 14, "Sector": 18, "Close": 9}
    for j, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(c, 11)
    # 指标说明 sheet
    ws2 = wb.create_sheet("指标说明")
    info = [
        ("Swing 宽表 — 指标说明", 14, True), (f"截至: {asof} | {len(df)} 只", 11, False), ("", 10, False),
        ("【① 超卖超买】", 11, True),
        ("  Close: 最新前复权收盘价", 10, False),
        ("  周J: 周KDJ的J值, 输出近4周swing极值。近4周内若到达极值(J<15或>95)取最近极值侧极值(翻转取最新侧); 无极值按当前方向(涨取min/跌取max)。始终输出一个值", 10, False),
        ("", 10, False),
        ("【② 趋势尾声 — 背离】", 11, True),
        ("  周KDJ背离/周MACD背离/周RSI背离: 价格新低(底背离)或新高(顶背离) vs 指标未新低/新高。周KDJ背离用J为指标且要求两极值点J<10(底)/J>90(顶)。过滤: 第二极值时效(周线近4周/日线近10日, 过期清除), 价差>2%, 间距≥4根", 10, False),
        ("    波动率取消: 价格从第二极值沿背离方向走 >6×20日收益标准差 → 视为已兑现→取消(银行低波~7%到位, 高波股~30%到位, 自适应)", 10, False),
        ("    反向取代: 若底/顶背离同时存在, 取第二极值更晚的那个(后来者取代前者)", 10, False),
        ("  日MACD背离/日RSI背离: 同上周线规则但日线级别, min_spacing=22, 无极值区约束", 10, False),
        ("  周度背离: 周KDJ/MACD/RSI任一背离 → +1底/-1顶/±1双向。周J超卖(<20)不显示顶背离; 超买(>80)不显示底背离", 10, False),
        ("  日度背离: 日KDJ/MACD/RSI任一背离 → +1底/-1顶/±1双向", 10, False),
        ("", 10, False),
        ("【② 趋势尾声 — DeMark TD Sequential】", 11, True),
        ("  周度DeMark / 日度DeMark: +9=买方setup完成(近8日/4周内, 非瞬时); +13=买方countdown完成(setup后数到13根 收盘<2根前低, 近4周/8日内完成, 底反转); -9/-13=卖方对称(顶反转)。正=买方, 负=卖方", 10, False),
        ("", 10, False),
        ("【③ 多空平衡】(1=命中, 空=未命中)", 11, True),
        ("  十字星(3d): 近3日十字星个数。十字星=实体/振幅≤10% 且 振幅/开盘≥0.5%(经典十字星), 或 实体绝对值≤0.02(价≥5)/≤0.01(价<5)(小实体)", 10, False),
        ("  涨放量跌缩量: 近10日≥2上涨日且≥1下跌日; 上涨日均量≥1.5×下跌日均量; 且每个下跌日成交量<当日20日量均(持续缩量)。多头承接信号", 10, False),
        ("  climax: 极端价格变动 + 放量 + 极端位置。+1=放量大涨在60日高位(最后一涨); -1=放量大跌在60日低位(最后一跌); ±1=双向; 取近5日最新", 10, False),
        ("    触发: |收盘-前收|≥4×ATR(14) 或 |日收益率|≥近250日99.5分位; 且 量≥1.5×20日量均; 位置=收盘在60日区间分位(高位≥0.85/低位≤0.15)", 10, False),
        ("    回测(4年/268只): climax bottom 20日forward +8.4%/胜率70.6%; climax top 20日forward正(动量延续, 预警较弱, 保留参考)", 10, False),
        ("", 10, False),
        ("【④ 企稳上行 — Cross】", 11, True),
        ("  KDJ Cross / 日MACD Cross / 5_10 Cross: 日线快线上穿慢线=金叉, 下穿=死叉", 10, False),
        ("    近2日金叉=1 / 死叉=-1; 按gap斜率线性推算将交叉=预1 / 预-1; 否则空", 10, False),
        ("    KDJ: K上穿D; MACD: DIF上穿DEA; 5_10: MA5上穿MA10", 10, False),
        ("    MACD 专属: 预告窗口=3日(其他2日), 且需过去3根MACD柱(DIF-DEA)依次单边且加速, 来回震荡不触发预告", 10, False),
    ]
    if cols and "MRS_HSI" in cols:  # HK 专属
        info += [
            ("", 10, False),
            ("【⑤ HK 专属 — 趋势/Mansfield RS】", 11, True),
            ("  年线斜率: MA200 近5日变化率 ×52 (年化, %)。正值=年线上行", 10, False),
            ("  MRS_HSI: Mansfield RS vs 2800.HK(盈富基金,HSI代理) = (个股/2800.HK)/(252日均线)-1 ×100。0=在均线上, 正=跑赢, 负=跑输", 10, False),
            ("  MRS_HSI_1W / MRS_HSI_4W: MRS_HSI 分数的 5日 / 20日 变化", 10, False),
            ("  MRS_HD: Mansfield RS vs 3110.HK(恒生高股息率ETF) = (个股/3110.HK)/(252日均线)-1 ×100。正=跑赢高股息指数", 10, False),
            ("  MRS_HD_1W / MRS_HD_4W: MRS_HD 分数的 5日 / 20日 变化", 10, False),
            ("  (HK股票池无sector, 不算 MRS_行业)", 10, False),
        ]
    elif cols and "MRS_SPY" in cols:  # 美股专属
        info += [
            ("", 10, False),
            ("【⑤ 美股专属 — 趋势/Mansfield RS】", 11, True),
            ("  年线斜率: MA200 近5日变化率 ×52 (年化, %)。正值=年线上行", 10, False),
            ("  MRS_SPY: Mansfield RS vs SPY = (个股/SPY)/(252日均线)-1 ×100。0=在均线上, 正=跑赢S&P, 负=跑输", 10, False),
            ("  MRS_行业: Mansfield RS vs 行业ETF (GICS sector→XLK/XLE/XLV...) 同法", 10, False),
            ("  MRS_SPY_1W / MRS_SPY_4W: MRS_SPY 分数的 5日 / 20日 变化", 10, False),
        ]
    for i, (t, s, b) in enumerate(info, 1):
        c = ws2.cell(i, 1, t); c.font = Font(size=s, bold=b, color="1F4E78" if b and s >= 12 else "000000")
    ws2.column_dimensions["A"].width = 110
    wb.save(out)
    print(f"[Excel] -> {out}")


def main():
    ap = argparse.ArgumentParser(description="Swing 宽表")
    ap.add_argument("--asof", default=None)
    args = ap.parse_args()
    cfg = yaml.safe_load(open(ROOT / "config.yaml", encoding="utf-8"))
    asof = args.asof or cfg["data"]["asof"] or dt.date.today().strftime("%Y-%m-%d")
    out_dir = ROOT / cfg["output"]["dir"]; out_dir.mkdir(exist_ok=True)

    pool = P.load_pool(cfg["data"]["pool_csv"])
    print(f"[Pool] {len(pool)} 只 | asof={asof}")
    codes = [c for c, _, _ in pool]
    # EODHD 并行预取 (第一数据源); 缺失/失败走 Wind 回退
    eodhd_bag = {}
    try:
        import eodhd
        eodhd_bag = eodhd.fetch_all_eodhd(codes, asof, cfg["data"]["lookback_days"], workers=5)
    except Exception as e:
        print(f"[EODHD] 预取失败, 全部走 Wind: {e}")
    if eodhd_bag:
        print(f"[EODHD] 命中 {len(eodhd_bag)}/{len(codes)}, 缺失走 Wind 回退")
    # HK 基准 (2800.HK 盈富基金) 用于 RS 指标
    import hk_rs
    hk_bench = {}
    try:
        hk_bench = hk_rs.fetch_benchmarks(asof, cfg["data"]["lookback_days"])
        print(f"[Bench] 2800.HK {'OK' if hk_bench else '缺失'}")
    except Exception as e:
        print(f"[Bench] 拉取失败: {e}")
    # Wind fetcher 懒构造 (仅 EODHD 缺失时才建; DB secret 缺失则跳过, 不崩)
    fetcher = [None]  # 用 list 闭包可变
    def get_fetcher():
        if fetcher[0] is None:
            try:
                fetcher[0] = P.make_fetcher(cfg["data"]["lookback_days"])
            except Exception as e:
                print(f"[Wind] 不可用 ({e}), 跳过 Wind 回退")
                fetcher[0] = False
        return fetcher[0] if fetcher[0] is not False else None
    rows = []
    t0 = time.time()
    n_wind_fb = 0
    for idx, (code, name, sector) in enumerate(pool):
        if (idx + 1) % 100 == 0:
            print(f"  [{idx+1}/{len(pool)}] {time.time()-t0:.0f}s ...")
        df = eodhd_bag.get(code)
        if df is None or df.empty:
            f = get_fetcher()
            if f is None:
                continue
            daily = P.fetch_daily(f, code, asof)   # Wind 回退
            n_wind_fb += 1
        else:
            daily = P.forward_adjust(df)
            daily = daily[daily["date"] <= pd.to_datetime(asof)].reset_index(drop=True)
        if daily is None or daily.empty or len(daily) < 60:
            continue
        try:
            r = sw.analyze(daily)
            if r:
                r.update(hk_rs.compute_extras(daily, hk_bench))
        except Exception as e:
            print(f"  [ERR] {code}: {e}"); r = None
        if r is None:
            continue
        r["Ticker"] = to_bloomberg(code)
        r["Name"] = name
        r["Sector"] = sector
        rows.append(r)
    if fetcher[0] and fetcher[0] is not False:
        fetcher[0].close()
    if eodhd_bag:
        print(f"[Wind回退] {n_wind_fb} 只")
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df[COL_ORDER_HK]
        # 排序: 有周度背离/DeMark/Cross 信号优先
        def _has_sig(r):
            return any(str(r.get(c)) not in ("", "nan", "None") for c in
                       ["周度背离", "日度背离", "周度DeMark", "日度DeMark",
                        "KDJ Cross", "日MACD Cross", "5_10 Cross"])
        df["_sig"] = df.apply(_has_sig, axis=1)
        df = df.sort_values(["_sig", "Ticker"], ascending=[False, True]).drop(columns="_sig").reset_index(drop=True)
    print(f"\n[完成] {len(df)} 只, 耗时 {time.time()-t0:.0f}s")
    if df.empty:
        print("无数据"); return
    out_path = out_dir / OUT_XLSX
    try:
        write_excel(df, asof, out_path, col_order=COL_ORDER_HK)
    except PermissionError:
        out_path = out_dir / f"swing_{asof}_alt.xlsx"
        print(f"[WARN] 主文件被占用, 写到 {out_path.name}")
        write_excel(df, asof, out_path, col_order=COL_ORDER_HK)
    # 生成网站页 (两 tab HK/US, 含下载按钮 + 表格) + xlsx 复制到根目录(供下载)
    try:
        import shutil
        shutil.copy(out_path, ROOT / OUT_XLSX)
        import gen_site
        gen_site.build(ROOT / OUT_XLSX, ROOT / "US_Swing_Pattern.xlsx", ROOT / "index.html")
        print(f"[Site] -> {ROOT / 'index.html'}")
    except Exception as e:
        print(f"[Site] 生成失败: {e}")


if __name__ == "__main__":
    main()
