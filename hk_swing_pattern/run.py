"""CLI 入口: 港股 背离+十字星+筹码峰 三层信号合成器

用法:
  python run.py                     # 截至今天
  python run.py --asof 2026-07-20
"""
from __future__ import annotations
import sys, io, os, argparse, datetime as dt
from pathlib import Path
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import yaml
import pandas as pd

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / "src"))
ETS = r"D:\equity-trend-screener\src"
if ETS not in sys.path:
    sys.path.insert(0, ETS)

import provider as P  # noqa: E402
import scan as scanmod  # noqa: E402
import chip as chipmod  # noqa: E402
import synth as sigmod  # noqa: E402
import frontend as fe  # noqa: E402


def load_cfg() -> dict:
    return yaml.safe_load(open(ROOT / "config.yaml", encoding="utf-8"))


def write_excel(df: pd.DataFrame, asof: str, out: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    tier_fill = {"高": "C6EFCE", "中": "FFF2CC", "低": "F0F0F0"}

    def sheet(ws, sub, title):
        ws.cell(1, 1, title).font = Font(bold=True, size=13, color="1F4E78")
        if sub.empty:
            ws.cell(3, 1, "(无)").font = Font(italic=True, color="888888"); return
        cols = list(sub.columns)
        for j, c in enumerate(cols, 1):
            cell = ws.cell(3, j, c); cell.fill = head_fill; cell.font = head_font
            cell.alignment = Alignment(horizontal="center")
        for i, (_, r) in enumerate(sub.iterrows()):
            rownum = 4 + i
            fill = PatternFill("solid", fgColor=tier_fill.get(r.get("Tier", "低"), "F0F0F0"))
            for j, c in enumerate(cols, 1):
                v = r[c]
                if isinstance(v, float) and pd.isna(v):
                    v = None
                cc = ws.cell(rownum, j, v); cc.fill = fill
        ws.freeze_panes = "A4"

    cols = ["Ticker", "Name", "Sector", "Side", "Tier", "Score",
            "DivType", "DivDate", "WeeksSinceDiv", "Anchor",
            "TrigDate", "TrigClose", "DojiLow", "DojiHigh",
            "周K", "周D", "周J", "40wJmin", "40wJmax",
            "ChipExp", "ChipTur", "ExpBand", "TurBand",
            "NearAnchor", "VolDiv", "QuietDoji", "JExtreme", "WkSignal", "DailySync",
            "VolRatio"]
    ws1 = wb.active; ws1.title = "底买"
    sheet(ws1, df[df["Side"] == "底买"][cols] if not df.empty else pd.DataFrame(), f"底买信号 — {asof}")
    ws2 = wb.create_sheet("顶卖")
    sheet(ws2, df[df["Side"] == "顶卖"][cols] if not df.empty else pd.DataFrame(), f"顶卖信号 — {asof}")
    ws3 = wb.create_sheet("全量")
    sheet(ws3, df[cols] if not df.empty else pd.DataFrame(), f"全量 — {asof}")
    wb.save(out)
    print(f"[Excel] -> {out}")


def write_html(df: pd.DataFrame, gmap: dict, cfg: dict, asof: str, out: Path):
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    ch = cfg["chip"]
    top = df[df["Tier"] == "高"].head(20) if not df.empty else pd.DataFrame()
    if top.empty:
        top = df.head(20) if not df.empty else top

    parts = ["<!DOCTYPE html><html lang='zh'><head><meta charset='utf-8'>",
             f"<title>背离+十字星+筹码峰 {asof}</title>",
             "<style>*{{margin:0;padding:0;box-sizing:border-box}}"
             "body{font-family:'Segoe UI','Microsoft YaHei',Arial;background:#f5f6fa;color:#2c3e50;padding:18px}"
             "h1{font-size:20px;color:#1F4E78}h2{font-size:15px;color:#1F4E78;margin:18px 0 6px}"
             ".meta{color:#666;font-size:12px;margin-bottom:8px}"
             ".card{background:#fff;border-radius:8px;padding:12px;margin-bottom:18px;box-shadow:0 1px 3px rgba(0,0,0,.08)}"
             ".tag{display:inline-block;padding:2px 8px;border-radius:4px;font-size:12px;font-weight:600;margin-right:6px}"
             ".t-high{background:#d5f5e3;color:#1e8449}.t-mid{background:#fdebd0;color:#b9770e}.t-low{background:#eee;color:#888}"
             ".t-buy{background:#e8f4ff;color:#1f6fb4}.t-sell{background:#fde8ef;color:#a12c5b}"
             "</style></head><body>",
             f"<h1>背离 + 十字星 + 筹码峰 三层信号</h1><div class='meta'>截至 {asof} | "
             f"高 {len(df[df['Tier']=='高']) if not df.empty else 0} / "
             f"中 {len(df[df['Tier']=='中']) if not df.empty else 0} / "
             f"低 {len(df[df['Tier']=='低']) if not df.empty else 0} | "
             f"底买 {len(df[df['Side']=='底买']) if not df.empty else 0} / "
             f"顶卖 {len(df[df['Side']=='顶卖']) if not df.empty else 0}</div>"]

    first = True
    for _, r in top.iterrows():
        code = r["Ticker"]
        daily = gmap.get(code)
        parts.append(f"<div class='card'><h2>{code} {r['Name']} "
                     f"<span class='tag t-{r['Side']=='顶卖' and 'sell' or 'buy'}'>{r['Side']}</span>"
                     f"<span class='tag t-{ {'高':'high','中':'mid','低':'low'}[r['Tier']] }'>{r['Tier']} {r['Score']}分</span></h2>"
                     f"<div class='meta'>背离:{r['DivType']} {r['DivDate']}({r['WeeksSinceDiv']}w前) | "
                     f"trigger doji:{r['TrigDate']} @ {r['TrigClose']} | 周J:{r['周J']} | "
                     f"筹码命中 指数:{r['ChipExp']} 换手:{r['ChipTur']} | "
                     f"band Exp:{r['ExpBand']} Tur:{r['TurBand']}</div>")
        if daily is None:
            parts.append("<div class='meta'>无日线数据</div></div>"); continue
        try:
            parts.append(_plot_one(daily, r, cfg, first))
            first = False
        except Exception as e:
            parts.append(f"<div class='meta'>绘图失败: {e}</div>")
        parts.append("</div>")

    parts.append("</body></html>")
    with open(out, "w", encoding="utf-8") as f:
        f.write("".join(parts))
    print(f"[HTML]  -> {out}")


def _plot_one(daily: pd.DataFrame, r, cfg: dict, first: bool = False) -> str:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    ch = cfg["chip"]; win = cfg["data"]["chip_window"]
    d = daily.tail(win)
    exp = chipmod.compute_density_exp(daily, tau=ch["exp_tau"], nbins=ch["nbins"], window=win)
    tur = chipmod.compute_density_turnover(daily, r.get("FloatShares"), nbins=ch["nbins"],
                                           window=win, cap=ch["turnover_cap"])
    fig = make_subplots(rows=2, cols=1, shared_xaxes=False, vertical_spacing=0.12,
                        row_heights=[0.6, 0.4],
                        subplot_titles=("价格 (含 trigger doji 与 筹码密集带)", "筹码密度 (蓝=指数衰减 橙=换手率衰减)"))
    fig.add_trace(go.Scatter(x=d["date"], y=d["fwd_close"], name="收盘",
                             line=dict(color="#2980b9", width=1.4)), row=1, col=1)
    # doji trigger 点
    td = pd.to_datetime(r["TrigDate"])
    fig.add_trace(go.Scatter(x=[td], y=[r["TrigClose"]], mode="markers",
                             marker=dict(size=12, color="#e44", symbol="x"),
                             name="doji trigger"), row=1, col=1)
    fig.add_hline(y=r["Anchor"], line_dash="dot", line_color="#888",
                  annotation_text=f"anchor {r['Anchor']}", row=1, col=1)
    # 筹码带 (用 exp bands 画矩形)
    if exp:
        bands = chipmod.find_bands(exp[0], exp[1], ch["peak_order"], ch["peak_band_ratio"],
                                   ch["sig_peak_ratio"], ch["smooth_win"])
        for b in bands:
            fig.add_hrect(y0=b[0], y1=b[1], fillcolor="#27ae60", opacity=0.10,
                          line_width=0, row=1, col=1)
        fig.add_trace(go.Scatter(x=exp[0], y=exp[1], fill="tozeroy", name="指数衰减",
                                 line=dict(color="#2980b9", width=1.2)), row=2, col=1)
    if tur:
        fig.add_trace(go.Scatter(x=tur[0], y=tur[1], fill="tozeroy", name="换手率衰减",
                                 line=dict(color="#e67e22", width=1.2)), row=2, col=1)
    fig.update_layout(height=520, margin=dict(l=50, r=20, t=50, b=30), showlegend=False,
                      plot_bgcolor="#fff")
    return fig.to_html(include_plotlyjs=("inline" if first else False),
                       full_html=False, div_id=f"plot_{r['Ticker']}")


def main():
    ap = argparse.ArgumentParser(description="背离+十字星+筹码峰 三层信号")
    ap.add_argument("--asof", default=None)
    args = ap.parse_args()
    cfg = load_cfg()
    asof = args.asof or cfg["data"]["asof"] or dt.date.today().strftime("%Y-%m-%d")
    out_dir = ROOT / cfg["output"]["dir"]
    out_dir.mkdir(exist_ok=True)

    fetcher = P.make_fetcher(cfg["data"]["lookback_days"])
    df, gmap, float_map = scanmod.run(asof, cfg, fetcher)
    fetcher.close()

    if df.empty:
        print("[结果] 无信号"); return
    for side in ("底买", "顶卖"):
        sub = df[df["Side"] == side]
        print(f"\n=== {side} ({len(sub)}) ===")
        show = ["Ticker", "Name", "Tier", "Score", "DivDate", "WeeksSinceDiv",
                "TrigDate", "TrigClose", "周J", "ChipExp", "ChipTur"]
        print(sub[show].to_string(index=False))

    write_excel(df, asof, out_dir / f"div_doji_chip_{asof}.xlsx")
    html = fe.build_page(df, gmap, cfg, asof, float_map)
    out_html = out_dir / f"div_doji_chip_{asof}.html"
    out_html.write_text(html, encoding="utf-8")
    print(f"[HTML]  -> {out_html}")


if __name__ == "__main__":
    main()
