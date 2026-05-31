"""v5.4 Excel export: 4 sheets with GICS sector row coloring.

Sheets: 1.进场候选 / 2.持有关注 / 3.警报 / README
Styling: dark blue header, sector-based row colors, freeze top row.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


GICS_COLORS = {
    'Information Technology': 'E8F0FE',
    'Financials': 'FFF3E0',
    'Health Care': 'E8F5E9',
    'Consumer Discretionary': 'FCE4EC',
    'Consumer Staples': 'FFF9C4',
    'Energy': 'FFF8E1',
    'Industrials': 'E0F2F1',
    'Materials': 'F3E5F5',
    'Real Estate': 'ECEFF1',
    'Communication Services': 'E3F2FD',
    'Utilities': 'FBE9E7',
    'Other': 'F5F5F5',
}

HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)


def _write_sheet(ws, headers, rows, sector_col_idx=None):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(rows, 2):
        sector = str(row_data[sector_col_idx]) if sector_col_idx is not None else ''
        fill_color = GICS_COLORS.get(sector, 'F5F5F5')
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = THIN_BORDER

    ws.freeze_panes = 'A2'

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50), min_col=col_idx, max_col=col_idx):
            for cell in row:
                max_len = max(max_len, len(str(cell.value or '')))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 28)


def export_excel(three_tabs: dict, output_path: str):
    wb = Workbook()

    # Sheet 1: 进场候选
    ws1 = wb.active
    ws1.title = '1.进场候选'
    headers = ['代码', '名称', 'GICS行业', '子行业', '状态', 'SOS', '收盘价', 'MA10', 'MA10斜率%', 'MA200斜率%']
    rows = []
    for item in three_tabs['tab1']['sos_section']:
        rows.append([
            item['ticker'], item.get('name_cn', ''), item.get('gics_sector', ''),
            item.get('sub_industry', ''), item['state'], item.get('sos_setup_recent', ''),
            item.get('last_close'), item.get('ma10', ''),
            item.get('ma10_slope_pct', ''), item.get('ma200_slope', ''),
        ])
    for item in three_tabs['tab1']['pool_section']:
        rows.append([
            item['ticker'], item.get('name_cn', ''), item.get('gics_sector', ''),
            item.get('sub_industry', ''), item['state'], '',
            item.get('last_close'), item.get('ma10', ''),
            item.get('ma10_slope_pct', ''), item.get('ma200_slope', ''),
        ])
    _write_sheet(ws1, headers, rows, sector_col_idx=2)

    # Sheet 2: 持有关注
    ws2 = wb.create_sheet('2.持有关注')
    headers2 = ['代码', '名称', 'GICS行业', '子行业', '显示等级', '子状态', '收盘价', 'MA10', 'MA10斜率%', 'MA60', 'MA200斜率%', '持仓天数']
    rows2 = []
    for item in three_tabs['tab2']:
        rows2.append([
            item['ticker'], item.get('name_cn', ''), item.get('gics_sector', ''),
            item.get('sub_industry', ''), item.get('display_tier', item['substate']),
            item['substate'], item.get('last_close'), item.get('ma10', ''),
            item.get('ma10_slope_pct', ''), item.get('ma60', ''),
            item.get('ma200_slope', ''), item.get('days_in_state', ''),
        ])
    _write_sheet(ws2, headers2, rows2, sector_col_idx=2)

    # Sheet 3: 警报
    ws3 = wb.create_sheet('3.警报')
    headers3 = ['代码', '名称', 'GICS行业', '警报类型', '回撤%', '天数', '收盘价']
    rows3 = []
    for item in three_tabs['tab3']:
        rows3.append([
            item['ticker'], item.get('name_cn', ''), item.get('gics_sector', ''),
            item['label'] + ' ' + item['emoji'],
            item.get('pullback_dd_pct', ''),
            item.get('days_in_pullback', item.get('days_in_state', '')),
            item.get('last_close', ''),
        ])
    _write_sheet(ws3, headers3, rows3, sector_col_idx=2)

    # Sheet 4: README
    ws4 = wb.create_sheet('README')
    readme_lines = [
        'Wyckoff v5.4 港股趋势扫描器',
        '',
        '三个标签页说明：',
        '1. 进场候选：SOS信号（SETUP_OK/ENTANGLED近5日触发）+ 候选池',
        '2. 持有关注：所有TRENDING状态，display_tier可能升级',
        '3. 警报：A·PULLBACK(刚进入回调) / B·EXIT(触发退出) / C·PULLBACK(回调>=10日)',
        '',
        '排序标准：GICS行业数量 → display_tier等级 → MA10 10日斜率',
        'display_tier升级规则：MID/EARLY + 近5日60日新高 → STRONG',
        'MA10斜率 = (MA10_today / MA10_{10日前} - 1) * 100',
        '',
        '颜色说明：行颜色按GICS行业区分',
        f'扫描日期：{three_tabs.get("date", "")}',
    ]
    for row_idx, line in enumerate(readme_lines, 1):
        ws4.cell(row=row_idx, column=1, value=line)
    ws4.column_dimensions['A'].width = 60

    wb.save(output_path)
