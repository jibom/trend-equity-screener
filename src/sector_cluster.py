"""
港股板块热度聚类分析

目标: 在 sector_map 的 300+ 港股池中, 用聚类找出「市场关注度最高 + 持续创出新高」的行为集群,
     再映射到 GICS 板块; 同时直接按板块聚合打分, 交叉验证。

特征 (每只股票, 标准化后聚类):
  新高 (60日 + 250日双窗口):
    pct_of_high_250   = 最新收盘 / 250日最高   (1.0 = 处于年内最高)
    pct_of_high_60    = 最新收盘 / 60日最高
    new_high_ratio_60 = 近60日中「当日收盘=60日新高」的天数占比
    new_high_ratio_250= 近250日中「当日收盘=250日新高」的天数占比
  关注度 (放大倍数 + 绝对排名 结合):
    amt_surge         = 近20日均成交额 / 过去250日均成交额   (相对自身跃升)
    amt_rank_pct      = 近5日均成交额在池内百分位排名         (绝对体量)
  动量:
    ret_20 / ret_60 / ret_120
  趋势:
    ma_aligned        = ma10>ma20>ma30>ma50>ma100 (自适应, 新股缺长均线按已有均线判定)

板块打分 composite = 0.25·z(新高广度250) + 0.15·z(新高广度60) + 0.15·z(量放倍数)
                     + 0.10·z(成交金额占比1日) + 0.20·z(成交金额占比5日) + 0.15·z(60日涨幅)

用法:
  python src/sector_cluster.py --asof 2026-06-20
"""
from __future__ import annotations
import os, sys, io, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import numpy as np
import pandas as pd
import pymysql
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.dirname(__file__))
from db_config import DB_CONFIG, SECTOR_MAP_FILE

ASOF_DEFAULT = "2026-06-20"
LOOKBACK_DAYS = 1200         # 日历日, 覆盖周线3年(~780交易日) + 日线热身
MIN_DAYS = 60                # 至少 60 交易日即可 (新股放宽, 特征按 min(历史,窗口) 自适应)
N_CLUSTERS = 6


def latest_asof() -> str:
    """DB 里港股最新交易日 (YYYY-MM-DD)。用于避免 asof=今天 但数据 T-1/T-2 造成日期与数据不符。"""
    conn = pymysql.connect(**DB_CONFIG)
    try:
        df = pd.read_sql("SELECT MAX(TRADE_DT) AS d FROM hkshareeodprices", conn)
    finally:
        conn.close()
    return pd.to_datetime(str(df.iloc[0, 0]), format="%Y%m%d").strftime("%Y-%m-%d")


def load_pool() -> pd.DataFrame:
    df = pd.read_csv(SECTOR_MAP_FILE)
    df = df.dropna(subset=["code"]).copy()
    df["code"] = df["code"].astype(str).str.zfill(4).str.upper()
    df["code"] = df["code"].apply(lambda x: x if x.endswith(".HK") else x + ".HK")
    df = df.drop_duplicates(subset="code")
    df["sub_industry"] = df["sub_industry"].fillna("未分类")
    return df[["code", "name_cn", "sub_industry", "gics_sector", "hs_sector"]].rename(
        columns={"code": "Ticker", "name_cn": "Name",
                 "sub_industry": "SubIndustry", "gics_sector": "Sector",
                 "hs_sector": "HSSector"})


def fetch_all(pool_codes: list[str], asof: str) -> pd.DataFrame:
    end = asof.replace("-", "")
    start = (pd.to_datetime(asof) - pd.Timedelta(days=LOOKBACK_DAYS)).strftime("%Y%m%d")
    codes_sql = ",".join(f"'{c}'" for c in pool_codes)
    sql = f"""
        SELECT S_INFO_WINDCODE AS code, TRADE_DT,
               S_DQ_CLOSE, S_DQ_ADJOPEN, S_DQ_ADJHIGH, S_DQ_ADJLOW, S_DQ_ADJCLOSE,
               S_DQ_VOLUME, S_DQ_AMOUNT
        FROM hkshareeodprices
        WHERE TRADE_DT BETWEEN %s AND %s
          AND S_INFO_WINDCODE IN ({codes_sql})
        ORDER BY S_INFO_WINDCODE, TRADE_DT
    """
    conn = pymysql.connect(**DB_CONFIG)
    try:
        raw = pd.read_sql(sql, conn, params=(start, end))
    finally:
        conn.close()
    raw = raw[raw["TRADE_DT"] <= end].copy()
    return raw


def forward_adjust_group(g: pd.DataFrame) -> pd.DataFrame | None:
    """组内前复权: 最新日 fwd_close == raw_close"""
    g = g.sort_values("TRADE_DT").reset_index(drop=True)
    latest_raw = g["S_DQ_CLOSE"].iloc[-1]
    latest_adj = g["S_DQ_ADJCLOSE"].iloc[-1]
    if pd.isna(latest_raw) or pd.isna(latest_adj) or latest_raw == 0:
        return None
    factor = latest_adj / latest_raw
    g["fwd_close"] = g["S_DQ_ADJCLOSE"] / factor
    g["fwd_high"] = g["S_DQ_ADJHIGH"] / factor
    g["fwd_low"] = g["S_DQ_ADJLOW"] / factor
    g["amt"] = g["S_DQ_AMOUNT"]
    g["vol"] = g["S_DQ_VOLUME"]
    return g


def compute_share_pctile200(raw: pd.DataFrame, pool: pd.DataFrame, group_col: str) -> pd.Series:
    """每个分组(细分行业/GICS板块) 的「5日成交金额占比」在过去200天的百分位(0-100)。

    衡量当前关注度相对【自身历史】的高低, 避免大市值板块(如互联网)长期占据综合分高位。
    返回 Series, 索引=分组名, 值=百分位(0-100)。
    """
    long = raw[["code", "TRADE_DT", "S_DQ_AMOUNT"]].merge(
        pool[["Ticker", group_col]], left_on="code", right_on="Ticker", how="inner")
    gd = long.groupby(["TRADE_DT", group_col])["S_DQ_AMOUNT"].sum().reset_index()
    td = long.groupby("TRADE_DT")["S_DQ_AMOUNT"].sum().reset_index().rename(columns={"S_DQ_AMOUNT": "tot"})
    gd = gd.merge(td, on="TRADE_DT").sort_values([group_col, "TRADE_DT"])
    gd["sub5"] = gd.groupby(group_col)["S_DQ_AMOUNT"].transform(lambda x: x.rolling(5, min_periods=5).sum())
    gd["tot5"] = gd.groupby(group_col)["tot"].transform(lambda x: x.rolling(5, min_periods=5).sum())
    gd["share5"] = np.where(gd["tot5"] > 0, gd["sub5"] / gd["tot5"], np.nan)

    def _pct(s):
        s = s.dropna()
        if len(s) < 10:
            return np.nan
        last = s.iloc[-1]
        hist = s.iloc[-200:]
        return float((hist <= last).mean() * 100)

    return gd.groupby(group_col)["share5"].apply(_pct)


def compute_share_rotation(raw: pd.DataFrame, pool: pd.DataFrame, group_col: str,
                           lookback: int = 10) -> pd.DataFrame:
    """「5日成交占比200日分位」在过去 `lookback` 天的变化 + 量能匹配 → 资金轮动信号。

    返回 DataFrame (按 delta10 降序), 列:
      group, n, pctile_now, pctile_ago, delta10, amt5_now_yi, amt5_chg_pct, flag
    flag:
      'lose' 高位失宠  = pctile_ago>=80 且 pctile_now>=60 且 delta10<=-15 且 5日均额<-10% (高位且仍高, 量能萎缩; 已跌到低位的不算)
      'gain' 低位放量突破 = pctile_ago<=40 且 pctile_now>=50 且 delta10>=30 且 5日均额>30% (从低位放量突破; 仅微幅流入的不算)
      ''    无
    """
    long = raw[["code", "TRADE_DT", "S_DQ_AMOUNT"]].merge(
        pool[["Ticker", group_col]], left_on="code", right_on="Ticker", how="inner")
    gd = long.groupby(["TRADE_DT", group_col])["S_DQ_AMOUNT"].sum().reset_index()
    td = long.groupby("TRADE_DT")["S_DQ_AMOUNT"].sum().reset_index().rename(columns={"S_DQ_AMOUNT": "tot"})
    gd = gd.merge(td, on="TRADE_DT").sort_values([group_col, "TRADE_DT"])
    gd["sub5"] = gd.groupby(group_col)["S_DQ_AMOUNT"].transform(lambda x: x.rolling(5, min_periods=5).sum())
    gd["tot5"] = gd.groupby(group_col)["tot"].transform(lambda x: x.rolling(5, min_periods=5).sum())
    gd["share5"] = np.where(gd["tot5"] > 0, gd["sub5"] / gd["tot5"], np.nan)
    nmap = pool.groupby(group_col)["Ticker"].count()

    rows = []
    for g, d in gd.groupby(group_col):
        d = d.reset_index(drop=True)
        s = d["share5"].values
        a = d["sub5"].values      # 5日累计成交额 (千元)
        L = len(s)
        if L < 60:
            continue

        def _pct(i):
            w = s[max(0, i - 199):i + 1]
            w = w[~np.isnan(w)]
            if len(w) < 10:
                return np.nan
            return float((w <= s[i]).mean() * 100)

        now_i, ago_i = L - 1, max(0, L - 1 - lookback)
        pnow, pago = _pct(now_i), _pct(ago_i)
        if np.isnan(pnow) or np.isnan(pago):
            continue
        amt_now, amt_ago = float(a[now_i]), float(a[ago_i])
        amt_chg = (amt_now / amt_ago - 1) if amt_ago > 0 else None
        delta = pnow - pago
        flag = ""
        if pago >= 80 and pnow >= 60 and delta <= -15 and (amt_chg or 0) < -0.10:
            flag = "lose"
        elif pago <= 40 and pnow >= 50 and delta >= 30 and (amt_chg or 0) > 0.30:
            flag = "gain"
        rows.append({group_col: g, "n": int(nmap.get(g, 0)),
                     "pctile_now": round(pnow, 1), "pctile_ago": round(pago, 1),
                     "delta10": round(delta, 1),
                     "amt5_now_yi": round(amt_now / 1e5, 2) if amt_now else None,
                     "amt5_chg_pct": round(amt_chg * 100, 1) if amt_chg is not None else None,
                     "flag": flag})
    res = pd.DataFrame(rows)
    if len(res):
        res = res.sort_values("delta10", ascending=False).reset_index(drop=True)
    return res


def compute_features(g: pd.DataFrame) -> dict | None:
    if g is None or len(g) < MIN_DAYS:
        return None
    c = g["fwd_close"].values
    amt = g["amt"].values
    n = len(g)
    close = float(c[-1])

    # 新高: 窗口取 min(历史, 窗口), 新股上市即在高点 → pct=1.0 (合理)
    w250 = min(n, 250)
    high_250 = np.max(c[-w250:])
    high_60 = np.max(c[-60:])
    high_126 = np.max(c[-min(n, 126):])
    pct_high_250 = close / high_250 if high_250 > 0 else np.nan
    pct_high_60 = close / high_60 if high_60 > 0 else np.nan
    pct_high_126 = close / high_126 if high_126 > 0 else np.nan

    s = pd.Series(c)
    # 60日新高天数占比
    roll60_max = s.rolling(60, min_periods=60).max()
    nh_60 = int((s[-60:] == roll60_max[-60:]).sum())
    nh_ratio_60 = nh_60 / 60.0
    # 126日(≈6月) 新高天数占比
    w126 = min(n, 126)
    roll126_max = s.rolling(w126, min_periods=w126).max()
    nh_126 = int((s[-w126:] == roll126_max[-w126:]).sum())
    nh_ratio_126 = nh_126 / float(w126)
    # 250日 (或全部历史) 新高天数占比
    roll_long = s.rolling(w250, min_periods=w250).max()
    nh_long = int((s[-w250:] == roll_long[-w250:]).sum())
    nh_ratio_250 = nh_long / float(w250)

    # 关注度: 量放倍数 (20日均额/250日均额, 相对自身放大)
    amt_20 = np.nanmean(amt[-20:])
    amt_long = np.nanmean(amt[-w250:])
    amt_surge = amt_20 / amt_long if amt_long and amt_long > 0 else np.nan

    # 成交金额: 最近1日 / 最近5日均值 (单位千元, 千港元)
    amt_1d = float(amt[-1]) if not np.isnan(amt[-1]) else np.nan
    amt_5d = float(np.nanmean(amt[-5:]))

    # 动量: 不足窗口的用上市以来涨幅代替 (避免 NaN)
    ret = lambda p: (close / c[-p - 1] - 1) if n > p and c[-p - 1] else None
    ret_20 = ret(20) if ret(20) is not None else np.nan
    ret_60 = ret(60) if ret(60) is not None else np.nan
    ret_120 = ret(120) if ret(120) is not None else (close / c[0] - 1)

    # 均线多头排列: 按"可计算的最长前缀"判定; 新股缺长均线(如250D)时,
    # 只要已有均线(>=4根)严格递减即算多头。链 10>20>30>50>100 (250D 不作硬性要求)
    ma_windows = (10, 20, 30, 50, 100)
    ma = {w: s.rolling(w).mean().iloc[-1] for w in ma_windows}
    avail = [w for w in ma_windows if not pd.isna(ma[w])]
    if len(avail) >= 4:
        aligned = int(all(ma[avail[i]] > ma[avail[i + 1]] for i in range(len(avail) - 1)))
    else:
        aligned = 0
    # 中长期多头排列: close>MA20>MA50>MA100>MA200 (允许短期 10/30 纠缠, 需≥200日)
    ma200 = s.rolling(200).mean().iloc[-1] if n >= 200 else np.nan
    ma_stack = int(not pd.isna(ma200) and not pd.isna(ma[100])
                   and close > ma[20] > ma[50] > ma[100] > ma200)

    return dict(
        pct_high_250=pct_high_250, pct_high_60=pct_high_60, pct_high_126=pct_high_126,
        nh_ratio_60=nh_ratio_60, nh_ratio_126=nh_ratio_126, nh_ratio_250=nh_ratio_250,
        amt_20=amt_20, amt_surge=amt_surge, amt_1d=amt_1d, amt_5d=amt_5d,
        ret_20=ret_20, ret_60=ret_60, ret_120=ret_120,
        ma_aligned=aligned, ma_stack=ma_stack, close=close, days=n,
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--asof", default=None, help="截止日期; 默认取 DB 最新交易日")
    ap.add_argument("--k", type=int, default=N_CLUSTERS)
    args = ap.parse_args()
    asof = args.asof or latest_asof()

    pool = load_pool()
    print(f"=== 港股板块热度聚类 (asof={asof}) ===\n")
    print(f"[Pool] {len(pool)} 只")

    raw = fetch_all(pool["Ticker"].tolist(), asof)
    print(f"[Fetch] {len(raw)} 行, {raw['code'].nunique()} 只股票")

    rows = []
    for code, g in raw.groupby("code"):
        g = forward_adjust_group(g)
        feat = compute_features(g)
        if feat is None:
            continue
        meta = pool[pool["Ticker"] == code].iloc[0]
        rows.append({"Ticker": code, "Name": meta["Name"], "Sector": meta["Sector"],
                     "SubIndustry": meta["SubIndustry"], "HSSector": meta["HSSector"], **feat})
    df = pd.DataFrame(rows).reset_index(drop=True)

    # 近5日均成交额池内百分位 (个股关注度排名, 用于 Part2/3)
    df["amt_rank_pct"] = df["amt_5d"].rank(pct=True)
    df = df.dropna(subset=["pct_high_250", "amt_surge", "ret_60"]).reset_index(drop=True)
    print(f"[Feat] 有效 {len(df)} 只 (历史>=60交易日, 新股放宽)")

    # ---- 聚类 ----
    from sklearn.preprocessing import StandardScaler
    from sklearn.cluster import KMeans
    feat_cols = ["pct_high_250", "pct_high_60", "nh_ratio_60", "nh_ratio_250",
                 "amt_surge", "amt_rank_pct", "ret_20", "ret_60", "ret_120", "ma_aligned"]
    X = StandardScaler().fit_transform(df[feat_cols].values)
    km = KMeans(n_clusters=args.k, n_init=20, random_state=42)
    df["cluster"] = km.fit_predict(X)

    # 集群热度分: 用新高 + 关注度的 z 均值
    for col in ["pct_high_250", "nh_ratio_60", "amt_surge", "amt_rank_pct"]:
        df[f"_z_{col}"] = (df[col] - df[col].mean()) / df[col].std()
    df["hotness"] = df[["_z_pct_high_250", "_z_nh_ratio_60",
                        "_z_amt_surge", "_z_amt_rank_pct"]].mean(axis=1)

    clu = df.groupby("cluster").agg(
        n=("Ticker", "size"),
        pct_high_250=("pct_high_250", "mean"),
        nh_ratio_60=("nh_ratio_60", "mean"),
        amt_surge=("amt_surge", "mean"),
        amt_rank_pct=("amt_rank_pct", "mean"),
        ret_60=("ret_60", "mean"),
        hotness=("hotness", "mean"),
    ).sort_values("hotness", ascending=False)
    hot_cluster = int(clu.index[0])

    # ============ 通用分组打分 (GICS板块 / 细分行业) ============
    # composite = 新高广度 + 相对量能放大 + 当期关注度(5日占比在过去200天的百分位) + 动量
    # (用相对自身历史的百分位, 避免大市值板块长期占高位; 横截面占比1日/5日仅作展示)
    SCORE_COLS = ["breadth_250%", "breadth_60%", "mean_amt_surge", "share5_pctile200", "mean_ret_60"]
    pool_amt_1d = df["amt_1d"].sum()    # 全池最近1日总成交金额 (千元)
    pool_amt_5d = df["amt_5d"].sum()    # 全池最近5日均值总成交金额 (千元)
    pctile_sub = compute_share_pctile200(raw, pool, "SubIndustry")
    pctile_sec = compute_share_pctile200(raw, pool, "Sector")

    def group_score(d):
        g1 = d["amt_1d"].sum()          # 千元
        g5 = d["amt_5d"].sum()          # 千元
        return pd.Series({
            "n": len(d),
            "mean_pct_high_250": d["pct_high_250"].mean(),
            "breadth_250%": (d["pct_high_250"] >= 0.95).mean() * 100,
            "breadth_60%": (d["pct_high_60"] >= 0.98).mean() * 100,
            "mean_amt_surge": d["amt_surge"].mean(),
            "mean_amt_rank": d["amt_rank_pct"].mean(),
            "总成交金额_1日_亿": round(g1 / 1e5, 2),       # 千元 → 亿港元
            "总成交金额_5日均值_亿": round(g5 / 1e5, 2),
            "占比_1日%": round(g1 / pool_amt_1d * 100, 2) if pool_amt_1d else 0.0,
            "占比_5日%": round(g5 / pool_amt_5d * 100, 2) if pool_amt_5d else 0.0,
            "mean_ret_60": d["ret_60"].mean(),
            "mean_hotness": d["hotness"].mean(),
        })

    def rank_group(col, min_n, pctile):
        g = df.groupby(col).apply(group_score, include_groups=False)
        g = g[g["n"] >= min_n].copy()
        g["share5_pctile200"] = g.index.map(pctile).astype(float).fillna(50.0)
        z = g.copy()
        for c in SCORE_COLS:
            std = z[c].std()
            z[c] = (z[c] - z[c].mean()) / std if std and std > 0 else 0.0
        g["composite"] = (0.25 * z["breadth_250%"] + 0.15 * z["breadth_60%"]
                          + 0.15 * z["mean_amt_surge"] + 0.30 * z["share5_pctile200"]
                          + 0.15 * z["mean_ret_60"])
        return g.sort_values("composite", ascending=False).round(3)

    sub_rank = rank_group("SubIndustry", min_n=3, pctile=pctile_sub)     # 细分行业 (n<3 不参与)
    sector_rank = rank_group("Sector", min_n=5, pctile=pctile_sec)        # GICS 板块
    # 最热细分行业 = composite 第1 (小样本 n=3,4 也纳入, n 在表中可见供判断)
    hottest_sub = sub_rank.index[0]

    # ============ Part 2: 最热个股 (ma_aligned 作为列保留, 不单列表) ============
    top_stocks = df.sort_values("hotness", ascending=False).head(30)

    # ============ Part 3: 最热细分行业内个股热度排序 ============
    hot_sub_stocks = df[df["SubIndustry"] == hottest_sub].sort_values("hotness", ascending=False)

    # ============ 输出 Excel ============
    out_dir = os.path.join(os.path.dirname(__file__), "..", "output")
    out_path = os.path.join(out_dir, f"hk_sector_hotspot_{asof}.xlsx")
    sheets = [
        ("1-细分行业热度排名", "Part1 细分行业热度排名 (composite 降序, 样本>=3)", sub_rank),
        ("1b-GICS板块热度排名", "Part1 GICS板块热度排名 (样本>=5)", sector_rank),
        ("2-最热个股Top30", "Part2 最热个股 (hotness 降序)", top_stocks),
        (f"3-最热行业[{hottest_sub}]个股", f"Part3 最热细分行业【{hottest_sub}】个股热度排序", hot_sub_stocks),
        ("集群概览", "KMeans 集群概览 (按热度降序, 参考)", clu.reset_index()),
        ("个股特征全量", "个股特征 + 聚类标签 (全量)", df.sort_values("hotness", ascending=False)),
    ]
    _write_excel(sheets, asof, out_path)
    print(f"\n[Excel] -> {out_path}")

    # ============ 终端输出: 三段式 ============
    pd.set_option("display.width", 230)
    pd.set_option("display.max_colwidth", 18)
    bar = "=" * 72

    print("\n" + bar)
    print(" Part 1  市场最热行业 (含细分行业)")
    print(bar)
    print("\n--- 细分行业热度排名 (top 15, 样本>=3) ---")
    print(sub_rank[["n", "breadth_250%", "breadth_60%", "mean_amt_surge",
                    "总成交金额_1日_亿", "总成交金额_5日均值_亿", "占比_1日%", "占比_5日%",
                    "mean_ret_60", "composite"]]
          .head(15).to_string())
    print(f"\n>>> 最热细分行业 = 【{hottest_sub}】  n={int(sub_rank.loc[hottest_sub,'n'])}  composite={sub_rank.loc[hottest_sub,'composite']}")
    print("\n--- GICS 板块热度排名 (样本>=5) ---")
    print(sector_rank[["n", "breadth_250%", "breadth_60%", "mean_amt_surge",
                       "总成交金额_1日_亿", "总成交金额_5日均值_亿", "占比_1日%", "占比_5日%",
                       "mean_ret_60", "composite"]]
          .to_string())

    print("\n" + bar)
    print(" Part 2  最热个股 (ma_aligned=1 为多头排列)")
    print(bar)
    c2 = ["Ticker", "Name", "SubIndustry", "close", "pct_high_250",
          "amt_surge", "amt_rank_pct", "ret_60", "ma_aligned", "hotness"]
    print("\n--- 热度 top 20 个股 ---")
    print(top_stocks[c2].head(20).round(3).to_string(index=False))

    print("\n" + bar)
    print(f" Part 3  最热细分行业【{hottest_sub}】个股热度排序")
    print(bar)
    c3 = ["Ticker", "Name", "close", "pct_high_250", "nh_ratio_60",
          "amt_surge", "amt_rank_pct", "ret_60", "ma_aligned", "hotness"]
    print(hot_sub_stocks[c3].round(3).to_string(index=False))


def _write_excel(sheets, asof, out_path):
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    title_font = Font(bold=True, size=13)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")

    def write_df(ws, title, d):
        ws.cell(row=1, column=1, value=f"{title} — {asof}").font = title_font
        d2 = d.reset_index() if (d.index.name or not isinstance(d.index, pd.RangeIndex)) else d
        for j, col in enumerate(d2.columns, 1):
            c = ws.cell(row=3, column=j, value=str(col))
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
        for i, row in enumerate(d2.itertuples(index=False), start=4):
            for j, v in enumerate(row, 1):
                ws.cell(row=i, column=j, value=(None if pd.isna(v) else v))

    for name, title, d in sheets:
        import re
        safe = re.sub(r"[\[\]\:*?/\\]", "", name)[:31]
        write_df(wb.create_sheet(safe), title, d)
    wb.save(out_path)


if __name__ == "__main__":
    main()
